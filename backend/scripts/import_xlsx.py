"""
Importa a planilha base RESUMO 3º CICLO.xlsx para o MongoDB.

Coleções criadas:
  - imoveis:       cadastro de imóveis (de QT01-QT33)
  - quarteiroes:   totais por quarteirão (de RG2CAB)
  - localidade:    informações da localidade (de QT01 e RG2CAB)

Uso:
  python import_xlsx.py /path/to/RESUMO.xlsx
  python import_xlsx.py --download  # baixa do servidor remoto

Idempotente: limpa as coleções antes de inserir.
"""
import os
import re
import sys
import argparse
import uuid
from pathlib import Path
from urllib.request import urlretrieve

import openpyxl
from pymongo import MongoClient
from dotenv import load_dotenv

ROOT = Path(__file__).resolve().parent.parent
load_dotenv(ROOT / ".env")

MONGO_URL = os.environ["MONGO_URL"]
DB_NAME = os.environ["DB_NAME"]

# Layout das colunas em cada bloco (lado par/ímpar) dentro de uma aba QT
LEFT_BLOCK = {
    "logradouro": "A",
    "lado_indicator": "E",  # número 1, 3, 5 ...
    "numero": "F",
    "seq": "G",
    "tipo": "H",
    "hab": "I",
    "cao": "J",
    "gato": "K",
}
RIGHT_BLOCK = {
    "logradouro": "M",
    "lado_indicator": "Q",  # número 2, 4, 6 ...
    "numero": "R",
    "seq": "S",
    "tipo": "T",
    "hab": "U",
    "cao": "V",
    "gato": "W",
}

VALID_TIPOS = {"R", "C", "TB", "O", "PE"}


def cell(ws, row, col_letter):
    return ws[f"{col_letter}{row}"].value


def is_header_row(ws, row):
    v = cell(ws, row, "A")
    return isinstance(v, str) and "Rua ou Logradouro" in v


def is_total_row(ws, row):
    v = cell(ws, row, "A")
    return isinstance(v, str) and v.strip() in {"Tipo", "Total"}


def clean_str(v):
    if v is None:
        return ""
    return str(v).strip()


def to_int(v):
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return None


def parse_qt_sheet(ws, quarteirao_num):
    """
    Itera o sheet QT, mantendo o 'lado' corrente para cada bloco.
    Retorna lista de imóveis ordenada conforme ordem original da planilha (caminhada do agente).
    """
    imoveis = []
    current_lado_left = None
    current_lado_right = None
    ordem_counter = 0

    in_block = False
    for row in range(1, ws.max_row + 1):
        # Header inicia um novo bloco
        if is_header_row(ws, row):
            in_block = False
            # próxima linha contém os lados
            next_row = row + 1
            current_lado_left = to_int(cell(ws, next_row, LEFT_BLOCK["lado_indicator"]))
            current_lado_right = to_int(cell(ws, next_row, RIGHT_BLOCK["lado_indicator"]))
            in_block = True
            continue

        if is_total_row(ws, row):
            in_block = False
            continue

        if not in_block:
            continue

        # Skip the row that only has lado indicators (right after header)
        only_lado = (
            cell(ws, row, LEFT_BLOCK["lado_indicator"]) is not None
            and cell(ws, row, LEFT_BLOCK["logradouro"]) is None
            and cell(ws, row, LEFT_BLOCK["numero"]) is None
            and cell(ws, row, LEFT_BLOCK["tipo"]) is None
        )
        if only_lado:
            continue

        # Extract left side
        for side_idx, (side_def, lado) in enumerate((
            (LEFT_BLOCK, current_lado_left),
            (RIGHT_BLOCK, current_lado_right),
        )):
            logradouro = clean_str(cell(ws, row, side_def["logradouro"]))
            numero = cell(ws, row, side_def["numero"])
            tipo = clean_str(cell(ws, row, side_def["tipo"]))
            if not (logradouro or numero or tipo):
                continue
            if tipo and tipo not in VALID_TIPOS:
                # ignore weird values
                tipo = ""
            ordem_counter += 1
            imoveis.append({
                "id": str(uuid.uuid4()),
                "quarteirao": str(quarteirao_num),
                "lado": str(lado) if lado is not None else "",
                "logradouro": logradouro,
                "numero": "" if numero is None else str(numero),
                "seq": clean_str(cell(ws, row, side_def["seq"])),
                "tipo_imovel": tipo,
                "hab": to_int(cell(ws, row, side_def["hab"])) or 0,
                "cao": to_int(cell(ws, row, side_def["cao"])) or 0,
                "gato": to_int(cell(ws, row, side_def["gato"])) or 0,
                "ordem": ordem_counter,
                "row_origem": row,
                "side_origem": side_idx,
            })

    return imoveis


def parse_rg2cab(ws):
    """RG2CAB: totais agregados por quarteirão."""
    rows = []
    # Row 1: title (ZONA 14...)
    # Row 2: headers
    # Row 3+: data
    headers_map = {
        "A": "quarteirao",
        "B": "residencia",
        "C": "comercio",
        "D": "outros",
        "E": "terreno_baldio",
        "F": "soma_predios",
        "G": "soma_imoveis",
        "H": "habitantes",
        "I": "cao",
        "J": "gato",
    }
    for row in range(3, ws.max_row + 1):
        qt = to_int(cell(ws, row, "A"))
        if qt is None:
            continue
        rec = {"id": str(uuid.uuid4()), "quarteirao": str(qt)}
        for col, key in headers_map.items():
            if key == "quarteirao":
                continue
            rec[key] = to_int(cell(ws, row, col)) or 0
        rows.append(rec)
    return rows


def get_localidade_info(wb):
    """Extrai info do cabeçalho de QT01."""
    ws = wb["QT01"]
    return {
        "id": str(uuid.uuid4()),
        "uf": clean_str(cell(ws, 3, "F")).replace("UF:", "").strip() or "RN",
        "municipio_codigo": clean_str(cell(ws, 4, "A")).replace("CÓDIGO:", "").strip(),
        "municipio_nome": clean_str(cell(ws, 4, "F")),
        "localidade_codigo": clean_str(cell(ws, 5, "A")).replace("CÓDIGO:", "").strip(),
        "localidade_nome": clean_str(cell(ws, 5, "F")),
        "zona": "14",
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", nargs="?", help="Caminho para o .xlsx")
    parser.add_argument("--download", action="store_true", help="Baixar do servidor")
    parser.add_argument(
        "--url",
        default="https://customer-assets.emergentagent.com/job_d19fdbd1-2e97-4f77-86d8-b0329101dd0d/artifacts/q5rv7i0m_RESUMO%203%C2%BA%20CICLO.xlsx",
    )
    args = parser.parse_args()

    if args.download or not args.xlsx:
        path = "/tmp/_resumo_d1.xlsx"
        print(f"Baixando planilha → {path}")
        urlretrieve(args.url, path)
        xlsx_path = path
    else:
        xlsx_path = args.xlsx

    print(f"Lendo {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    client = MongoClient(MONGO_URL)
    db = client[DB_NAME]

    # Localidade
    loc = get_localidade_info(wb)
    db.localidade.delete_many({})
    db.localidade.insert_one(loc)
    print(f"✓ Localidade: {loc['municipio_nome']} / {loc['localidade_nome']}")

    # Imóveis (QT01..QT33)
    qt_sheets = sorted([s for s in wb.sheetnames if re.match(r"^QT\d+$", s)])
    print(f"Encontrados {len(qt_sheets)} quarteirões: {qt_sheets[0]}..{qt_sheets[-1]}")
    db.imoveis.delete_many({})
    all_imoveis = []
    for s in qt_sheets:
        qt_num = int(re.search(r"\d+", s).group())
        items = parse_qt_sheet(wb[s], qt_num)
        all_imoveis.extend(items)
        print(f"  {s} → {len(items)} imóveis")
    if all_imoveis:
        db.imoveis.insert_many(all_imoveis)
    print(f"✓ Total imóveis importados: {len(all_imoveis)}")

    # Quarteirões (RG2CAB)
    rg_rows = parse_rg2cab(wb["RG2CAB"])
    db.quarteiroes.delete_many({})
    if rg_rows:
        db.quarteiroes.insert_many(rg_rows)
    print(f"✓ Quarteirões (totais): {len(rg_rows)}")

    # Indexes
    db.imoveis.create_index("quarteirao")
    db.imoveis.create_index([("quarteirao", 1), ("lado", 1)])
    db.quarteiroes.create_index("quarteirao", unique=True)

    print("✅ Importação concluída.")
    client.close()


if __name__ == "__main__":
    main()
